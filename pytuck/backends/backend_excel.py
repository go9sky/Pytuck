"""
Pytuck Excel存储引擎

使用单个Excel工作簿（.xlsx），每个表一个工作表，可视化友好
"""

import json
import base64
from pathlib import Path
from typing import Any, Dict, List, Union, TYPE_CHECKING, Tuple, Optional
from datetime import datetime
from .base import StorageBackend
from ..common.exceptions import SerializationError
from .versions import get_format_version
from ..core.types import TypeRegistry

from ..common.options import ExcelBackendOptions

if TYPE_CHECKING:
    from ..core.storage import Table
    from openpyxl import Workbook


class ExcelBackend(StorageBackend):
    """Excel format storage engine (requires openpyxl)"""

    ENGINE_NAME = 'excel'
    REQUIRED_DEPENDENCIES = ['openpyxl']
    FORMAT_VERSION = get_format_version('excel')

    def __init__(self, file_path: Union[str, Path], options: ExcelBackendOptions):
        """
        初始化 Excel 后端

        Args:
            file_path: Excel 文件路径
            options: Excel 后端配置选项
        """
        assert isinstance(options, ExcelBackendOptions), "options must be an instance of ExcelBackendOptions"
        super().__init__(file_path, options)
        # 类型安全：将 options 转为具体的 ExcelBackendOptions 类型
        self.options: ExcelBackendOptions = options

    def save(self, tables: Dict[str, 'Table']) -> None:
        """保存所有表数据到Excel工作簿"""
        if self.options.read_only:
            raise SerializationError("Excel backend does not support read-only mode")
        try:
            from openpyxl import Workbook
        except ImportError:
            raise SerializationError("openpyxl is required for Excel backend. Install with: pip install pytuck[excel]")

        temp_path = self.file_path.parent / (self.file_path.name + '.tmp')
        try:
            wb = Workbook()
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 创建元数据工作表
            metadata_sheet = wb.create_sheet('_metadata', 0)
            metadata_sheet.append(['Key', 'Value'])
            metadata_sheet.append(['format_version', self.FORMAT_VERSION])
            metadata_sheet.append(['timestamp', datetime.now().isoformat()])
            metadata_sheet.append(['table_count', len(tables)])

            # 创建统一的表结构工作表 _pytuck_tables
            tables_sheet = wb.create_sheet('_pytuck_tables', 1)
            tables_sheet.append(['table_name', 'primary_key', 'next_id', 'comment', 'columns'])
            for table_name, table in tables.items():
                columns_json = json.dumps([
                    {
                        'name': col.name,
                        'type': col.col_type.__name__,
                        'nullable': col.nullable,
                        'primary_key': col.primary_key,
                        'index': col.index,
                        'comment': col.comment
                    }
                    for col in table.columns.values()
                ])
                tables_sheet.append([table_name, table.primary_key, table.next_id, table.comment or '', columns_json])

            # 根据配置隐藏元数据工作表
            if self.options.hide_metadata_sheets:
                metadata_sheet.sheet_state = 'hidden'
                tables_sheet.sheet_state = 'hidden'

            # 为每个表创建数据工作表
            for table_name, table in tables.items():
                self._save_table_to_workbook(wb, table_name, table)

            # 原子性保存
            wb.save(str(temp_path))

            if self.file_path.exists():
                self.file_path.unlink()
            temp_path.replace(self.file_path)

        except Exception as e:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except FileNotFoundError:
                    pass
            raise SerializationError(f"Failed to save Excel file: {e}")

    def load(self) -> Dict[str, 'Table']:
        """从Excel工作簿加载所有表数据"""
        if not self.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SerializationError("openpyxl is required for Excel backend. Install with: pip install pytuck[excel]")

        try:
            wb = load_workbook(
                filename=str(self.file_path), read_only=self.options.read_only, data_only=True, keep_links=False
            )

            # 从 _pytuck_tables 工作表读取所有表的 schema
            tables_schema: Dict[str, Dict[str, Any]] = {}
            if '_pytuck_tables' in wb.sheetnames:
                tables_sheet = wb['_pytuck_tables']
                rows = list(tables_sheet.iter_rows(min_row=2, values_only=True))
                for row in rows:
                    if row[0]:  # table_name 不为空
                        table_name = row[0]
                        tables_schema[table_name] = {
                            'primary_key': row[1],
                            'next_id': int(row[2]) if row[2] else 1,
                            'comment': row[3] if row[3] else None,
                            'columns': json.loads(row[4]) if row[4] else []
                        }

            # 获取所有数据表名（排除元数据表）
            table_names = [
                name for name in wb.sheetnames
                if not name.startswith('_')
            ]

            tables = {}
            for table_name in table_names:
                schema = tables_schema.get(table_name, {})
                table = self._load_table_from_workbook(wb, table_name, schema)
                tables[table_name] = table

            return tables

        except Exception as e:
            raise SerializationError(f"Failed to load Excel file: {e}")

    def exists(self) -> bool:
        """检查文件是否存在"""
        return self.file_path.exists()

    def delete(self) -> None:
        """删除文件"""
        if self.exists():
            self.file_path.unlink()

    def _save_table_to_workbook(self, wb: 'Workbook', table_name: str, table: 'Table') -> None:
        """保存单个表的数据到工作簿"""
        # 数据工作表
        data_sheet = wb.create_sheet(table_name)

        # 写入表头
        columns = list(table.columns.keys())
        # 确保主键列在列表中
        if table.primary_key not in columns:
            columns.insert(0, table.primary_key)

        # 如果启用了 persist_row_number 且 mapping='field'，确保行号字段在列表中
        persist_row_number = (
            self.options.row_number_mapping == 'field' and
            self.options.persist_row_number
        )
        row_number_field: Optional[str] = self.options.row_number_field_name if persist_row_number else None

        if persist_row_number and row_number_field is not None and row_number_field not in columns:
            columns.append(row_number_field)

        data_sheet.append(columns)

        # 写入数据行
        for excel_row, record in enumerate(table.data.values(), start=2):
            row: List[Any] = []
            for col_name in columns:
                # 如果是行号字段且启用了持久化，写入当前行号
                if persist_row_number and col_name == row_number_field:
                    row.append(excel_row)
                    continue

                value = record.get(col_name)
                column = table.columns.get(col_name)

                if value is None:
                    row.append('')
                elif column and column.col_type == bool:
                    # Excel 特殊处理：bool 转字符串 'TRUE'/'FALSE'
                    row.append('TRUE' if value else 'FALSE')
                elif column:
                    # 使用 TypeRegistry 统一序列化
                    row.append(TypeRegistry.serialize_for_text(value, column.col_type))
                else:
                    row.append(value)
            data_sheet.append(row)

    def _load_table_from_workbook(
        self, wb: 'Workbook', table_name: str, schema: Dict[str, Any]
    ) -> 'Table':
        """从工作簿加载单个表"""
        from ..core.storage import Table
        from ..core.orm import Column
        from datetime import datetime, date, timedelta
        from ..common.exceptions import SchemaError

        primary_key = schema.get('primary_key', 'id')
        next_id = schema.get('next_id', 1)
        table_comment = schema.get('comment')
        columns_data = schema.get('columns', [])

        # 判断是否应用行号映射
        # 只有在无 schema（外部 Excel）或显式设置 override 时才应用
        mapping_allowed = (self.options.row_number_mapping is not None) and (
            (not columns_data) or self.options.row_number_override
        )

        # 重建列
        columns = []

        if columns_data:
            # 有 schema（Pytuck 创建的文件）
            for col_data in columns_data:
                col_type = TypeRegistry.get_type_by_name(col_data['type'])
                column = Column(
                    col_data['name'],
                    col_type,
                    nullable=col_data['nullable'],
                    primary_key=col_data['primary_key'],
                    index=col_data.get('index', False),
                    comment=col_data.get('comment')
                )
                columns.append(column)
        else:
            # 无 schema（外部 Excel），从 headers 构建列
            data_sheet = wb[table_name]
            rows_preview = list(data_sheet.iter_rows(values_only=True, max_row=1))
            if rows_preview:
                headers = [h for h in rows_preview[0] if h]
                for name in headers:
                    columns.append(Column(name, str, nullable=True, primary_key=False))

        # 应用行号映射 - 处理列定义
        if mapping_allowed:
            mapping = self.options.row_number_mapping
            existing_col_names = [c.name for c in columns]

            if mapping == 'as_pk':
                # 确保主键列存在
                if primary_key not in existing_col_names:
                    pk_col = Column(primary_key, int, nullable=False, primary_key=True)
                    columns.insert(0, pk_col)
                else:
                    # 主键列已存在，检查类型
                    existing_col = next(c for c in columns if c.name == primary_key)
                    if existing_col.col_type != int:
                        raise SchemaError(
                            f"Cannot use row number as primary key: column '{primary_key}' is not int type",
                            details={'column_name': primary_key}
                        )
                    # 标记为主键
                    existing_col.primary_key = True

            elif mapping == 'field':
                field_name = self.options.row_number_field_name
                # 只有字段不存在时才添加
                if field_name not in existing_col_names:
                    columns.append(Column(field_name, int, nullable=True, primary_key=False))

        # 确保有主键列（对于外部 Excel）
        if not columns_data and not any(c.primary_key for c in columns):
            # 如果没有主键列，创建 id 列
            if primary_key not in [c.name for c in columns]:
                pk_col = Column(primary_key, int, nullable=False, primary_key=True)
                columns.insert(0, pk_col)
            else:
                # 将已有的同名列标记为主键
                for c in columns:
                    if c.name == primary_key:
                        c.primary_key = True
                        break

        # 创建表
        table = Table(table_name, columns, primary_key, comment=table_comment)
        table.next_id = next_id

        # 读取数据
        data_sheet = wb[table_name]
        rows = list(data_sheet.iter_rows(values_only=True))

        max_int_pk = 0  # 用于更新 next_id

        if len(rows) > 1:
            headers = rows[0]
            # 使用 enumerate 获取 Excel 行号（数据行从第 2 行开始）
            for excel_row, row_data in enumerate(rows[1:], start=2):
                record = {}
                for col_name, value in zip(headers, row_data):
                    if col_name not in table.columns:
                        continue

                    column = table.columns[col_name]

                    # 处理空值
                    if value == '' or value is None:
                        value = None
                    elif column.col_type == bool:
                        # Excel 的 bool 特殊处理
                        if isinstance(value, bool):
                            pass  # 保持原样
                        elif isinstance(value, str):
                            value = (value.upper() == 'TRUE')
                        else:
                            value = bool(value)
                    elif column.col_type == bytes:
                        # bytes 需要特殊处理（base64 解码）
                        if value:
                            value = base64.b64decode(value)
                    elif column.col_type in (datetime, date, timedelta, list, dict, int, float):
                        # 使用 TypeRegistry 统一反序列化
                        value = TypeRegistry.deserialize_from_text(value, column.col_type)

                    record[col_name] = value

                # 应用行号映射 - 处理数据
                pk: Any = None
                if mapping_allowed:
                    mapping = self.options.row_number_mapping

                    if mapping == 'as_pk':
                        # 将行号作为主键值
                        record[primary_key] = excel_row
                        pk = excel_row

                    elif mapping == 'field':
                        field_name = self.options.row_number_field_name
                        # 如果字段已有非空值且 override=False，则不覆盖
                        if self.options.row_number_override or record.get(field_name) is None:
                            record[field_name] = excel_row
                        # 主键仍使用原有逻辑
                        pk = record.get(primary_key)
                        if pk is None:
                            # 如果没有主键值，使用自增
                            pk = table.next_id
                            table.next_id += 1
                            record[primary_key] = pk
                else:
                    # 不使用行号映射
                    pk = record.get(primary_key)
                    if pk is None and not columns_data:
                        # 外部 Excel 没有主键值，使用自增
                        pk = table.next_id
                        table.next_id += 1
                        record[primary_key] = pk

                table.data[pk] = record

                # 跟踪最大 int pk
                if isinstance(pk, int) and pk > max_int_pk:
                    max_int_pk = pk

        # 更新 next_id（如果主键是 int 类型）
        if max_int_pk >= table.next_id:
            table.next_id = max_int_pk + 1

        # 重建索引
        for col_name, column in table.columns.items():
            if column.index:
                if col_name in table.indexes:
                    del table.indexes[col_name]
                table.build_index(col_name)

        return table

    def get_metadata(self) -> Dict[str, Any]:
        """获取元数据"""
        if not self.exists():
            return {}

        try:
            file_stat = self.file_path.stat()
            file_size = file_stat.st_size
            modified_time = file_stat.st_mtime

            from openpyxl import load_workbook
            wb = load_workbook(str(self.file_path), read_only=True)

            metadata = {
                'engine': 'excel',
                'file_size': file_size,
                'modified': modified_time
            }

            # 尝试读取元数据工作表
            if '_metadata' in wb.sheetnames:
                sheet = wb['_metadata']
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        metadata[row[0]] = row[1]

            wb.close()
            return metadata

        except:
            return {}

    @classmethod
    def probe(cls, file_path: Union[str, Path]) -> Tuple[bool, Optional[Dict[str, Any]]]:
        """
        轻量探测文件是否为 Excel 引擎格式

        通过检查 Excel 文件（实际上是 ZIP）是否包含 _pytuck_tables 工作表来识别。
        使用 ZIP 方式而非 openpyxl 以避免依赖问题和提高性能。

        Returns:
            Tuple[bool, Optional[Dict]]: (是否匹配, 元数据信息或None)
        """
        try:
            file_path = Path(file_path).expanduser()
            if not file_path.exists():
                return False, {'error': 'file_not_found'}

            # 获取文件信息
            file_stat = file_path.stat()
            file_size = file_stat.st_size

            # 空文件不可能是有效的 Excel
            if file_size == 0:
                return False, {'error': 'empty_file'}

            # Excel 文件实际上是 ZIP 格式，先检查是否为 ZIP
            import zipfile
            if not zipfile.is_zipfile(file_path):
                return False, None

            try:
                with zipfile.ZipFile(str(file_path), 'r') as zf:
                    namelist = zf.namelist()

                    # 检查是否为 Excel 文件结构
                    if 'xl/workbook.xml' not in namelist:
                        return False, None

                    # 检查是否包含 _pytuck_tables 工作表的 XML 文件
                    # Excel 工作表在 ZIP 中存储为 xl/worksheets/sheetN.xml
                    # 工作表名称映射在 xl/workbook.xml 中
                    pytuck_tables_found = False

                    try:
                        # 读取 workbook.xml 来查找工作表名称
                        with zf.open('xl/workbook.xml') as f:
                            workbook_xml = f.read(8192).decode('utf-8')  # 只读前 8KB

                        # 简单的字符串检查，查找 _pytuck_tables 工作表
                        if '_pytuck_tables' in workbook_xml:
                            pytuck_tables_found = True

                    except (KeyError, UnicodeDecodeError):
                        pass

                    if not pytuck_tables_found:
                        return False, None

                    # 尝试获取更多元数据信息
                    format_version = None
                    table_count = None
                    timestamp = None

                    # 如果有 openpyxl 依赖，尝试获取更详细信息
                    if cls.is_available():
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook(str(file_path), read_only=True, data_only=True)

                            # 从 _metadata 工作表读取信息
                            if '_metadata' in wb.sheetnames:
                                metadata_sheet = wb['_metadata']
                                for row in metadata_sheet.iter_rows(min_row=2, values_only=True):
                                    if row[0] and row[1]:
                                        if row[0] == 'format_version':
                                            format_version = row[1]
                                        elif row[0] == 'timestamp':
                                            timestamp = row[1]
                                        elif row[0] == 'table_count':
                                            table_count = row[1]

                            wb.close()
                        except Exception:
                            # 如果 openpyxl 读取失败，仍然可以确认是 Pytuck Excel 格式
                            pass

                    # 成功识别为 Excel 格式
                    return True, {
                        'engine': 'excel',
                        'format_version': format_version,
                        'table_count': table_count,
                        'file_size': file_size,
                        'modified': file_stat.st_mtime,
                        'timestamp': timestamp,
                        'confidence': 'high' if cls.is_available() else 'medium'
                    }

            except zipfile.BadZipFile:
                return False, {'error': 'corrupted_excel_file'}

        except Exception as e:
            return False, {'error': f'probe_exception: {str(e)}'}
