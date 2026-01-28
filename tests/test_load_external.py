"""
测试 load_table 函数

测试从外部 CSV/Excel 文件加载数据到模型对象列表
"""

import tempfile
from pathlib import Path
from typing import Type
import pytest

from pytuck import Storage, declarative_base, Column
from pytuck.core.orm import PureBaseModel
from pytuck.tools import load_table
from pytuck.common.exceptions import ValidationError


class TestLoadTableCSV:
    """测试 CSV 文件加载"""

    def test_load_basic_csv(self, tmp_path: Path) -> None:
        """测试基本 CSV 加载"""
        # 创建测试 CSV 文件
        csv_file = tmp_path / "users.csv"
        csv_file.write_text("id,name,age\n1,Alice,20\n2,Bob,25\n3,Charlie,30\n", encoding='utf-8')

        # 创建模型
        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class User(Base):
            __tablename__ = 'users'
            id = Column(int, primary_key=True)
            name = Column(str)
            age = Column(int)

        # 加载数据
        users = load_table(User, str(csv_file))

        # 验证
        assert len(users) == 3
        assert users[0].id == 1
        assert users[0].name == 'Alice'
        assert users[0].age == 20
        assert users[1].id == 2
        assert users[1].name == 'Bob'
        assert users[1].age == 25

    def test_type_conversion_int_float(self, tmp_path: Path) -> None:
        """测试 int 和 float 类型转换"""
        csv_file = tmp_path / "data.csv"
        csv_file.write_text("id,score,count\n1,95.5,100\n2,88.0,200\n", encoding='utf-8')

        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class Data(Base):
            __tablename__ = 'data'
            id = Column(int, primary_key=True)
            score = Column(float)
            count = Column(int)

        data = load_table(Data, str(csv_file))

        assert len(data) == 2
        assert data[0].id == 1
        assert data[0].score == 95.5
        assert data[0].count == 100
        assert isinstance(data[0].score, float)
        assert isinstance(data[0].count, int)

    def test_type_conversion_bool(self, tmp_path: Path) -> None:
        """测试 bool 类型转换"""
        csv_file = tmp_path / "flags.csv"
        csv_file.write_text("id,active,verified\n1,true,1\n2,false,0\n3,yes,no\n", encoding='utf-8')

        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class Flag(Base):
            __tablename__ = 'flags'
            id = Column(int, primary_key=True)
            active = Column(bool)
            verified = Column(bool)

        flags = load_table(Flag, str(csv_file))

        assert len(flags) == 3
        assert flags[0].active is True
        assert flags[0].verified is True
        assert flags[1].active is False
        assert flags[1].verified is False
        assert flags[2].active is True
        assert flags[2].verified is False

    def test_nullable_column(self, tmp_path: Path) -> None:
        """测试可空列"""
        csv_file = tmp_path / "users.csv"
        csv_file.write_text("id,name,email\n1,Alice,alice@example.com\n2,Bob,\n", encoding='utf-8')

        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class User(Base):
            __tablename__ = 'users'
            id = Column(int, primary_key=True)
            name = Column(str)
            email = Column(str, nullable=True)

        users = load_table(User, str(csv_file))

        assert len(users) == 2
        assert users[0].email == 'alice@example.com'
        assert users[1].email is None

    def test_type_conversion_error(self, tmp_path: Path) -> None:
        """测试类型转换失败"""
        csv_file = tmp_path / "bad.csv"
        csv_file.write_text("id,age\n1,twenty\n", encoding='utf-8')

        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class User(Base):
            __tablename__ = 'users'
            id = Column(int, primary_key=True)
            age = Column(int)

        with pytest.raises(ValueError) as exc_info:
            load_table(User, str(csv_file))

        assert "row 2" in str(exc_info.value)

    def test_file_not_found(self) -> None:
        """测试文件不存在"""
        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class User(Base):
            __tablename__ = 'users'
            id = Column(int, primary_key=True)

        with pytest.raises(FileNotFoundError):
            load_table(User, '/nonexistent/file.csv')

    def test_unsupported_file_type(self, tmp_path: Path) -> None:
        """测试不支持的文件类型"""
        txt_file = tmp_path / "data.txt"
        txt_file.write_text("id,name\n1,Alice\n", encoding='utf-8')

        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class User(Base):
            __tablename__ = 'users'
            id = Column(int, primary_key=True)

        with pytest.raises(ValueError) as exc_info:
            load_table(User, str(txt_file))

        assert "Unsupported file type" in str(exc_info.value)

    def test_custom_delimiter(self, tmp_path: Path) -> None:
        """测试自定义分隔符"""
        csv_file = tmp_path / "data.csv"
        csv_file.write_text("id;name;age\n1;Alice;20\n2;Bob;25\n", encoding='utf-8')

        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class User(Base):
            __tablename__ = 'users'
            id = Column(int, primary_key=True)
            name = Column(str)
            age = Column(int)

        users = load_table(User, str(csv_file), delimiter=';')

        assert len(users) == 2
        assert users[0].name == 'Alice'

    def test_extra_columns_in_file(self, tmp_path: Path) -> None:
        """测试文件中有多余的列（应该被忽略）"""
        csv_file = tmp_path / "users.csv"
        csv_file.write_text("id,name,extra_col,age\n1,Alice,ignored,20\n", encoding='utf-8')

        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class User(Base):
            __tablename__ = 'users'
            id = Column(int, primary_key=True)
            name = Column(str)
            age = Column(int)

        users = load_table(User, str(csv_file))

        assert len(users) == 1
        assert users[0].name == 'Alice'
        assert users[0].age == 20
        # extra_col 不应该在对象上
        assert not hasattr(users[0], 'extra_col') or getattr(users[0], 'extra_col', None) is None


class TestLoadTableExcel:
    """测试 Excel 文件加载"""

    @pytest.fixture
    def has_openpyxl(self) -> bool:
        """检查是否安装了 openpyxl"""
        try:
            import openpyxl
            return True
        except ImportError:
            return False

    def test_load_basic_excel(self, tmp_path: Path, has_openpyxl: bool) -> None:
        """测试基本 Excel 加载"""
        if not has_openpyxl:
            pytest.skip("openpyxl not installed")

        from openpyxl import Workbook

        # 创建测试 Excel 文件
        xlsx_file = tmp_path / "users.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['id', 'name', 'age'])
        ws.append([1, 'Alice', 20])
        ws.append([2, 'Bob', 25])
        wb.save(str(xlsx_file))

        # 创建模型
        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class User(Base):
            __tablename__ = 'users'
            id = Column(int, primary_key=True)
            name = Column(str)
            age = Column(int)

        # 加载数据
        users = load_table(User, str(xlsx_file))

        # 验证
        assert len(users) == 2
        assert users[0].id == 1
        assert users[0].name == 'Alice'
        assert users[0].age == 20

    def test_load_specific_sheet(self, tmp_path: Path, has_openpyxl: bool) -> None:
        """测试加载指定工作表"""
        if not has_openpyxl:
            pytest.skip("openpyxl not installed")

        from openpyxl import Workbook

        # 创建测试 Excel 文件（多个工作表）
        xlsx_file = tmp_path / "data.xlsx"
        wb = Workbook()

        # 第一个工作表
        ws1 = wb.active
        ws1.title = 'Sheet1'
        ws1.append(['id', 'value'])
        ws1.append([1, 'first'])

        # 第二个工作表
        ws2 = wb.create_sheet('Sheet2')
        ws2.append(['id', 'value'])
        ws2.append([2, 'second'])

        wb.save(str(xlsx_file))

        # 创建模型
        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class Data(Base):
            __tablename__ = 'data'
            id = Column(int, primary_key=True)
            value = Column(str)

        # 加载 Sheet2
        data = load_table(Data, str(xlsx_file), sheet_name='Sheet2')

        assert len(data) == 1
        assert data[0].id == 2
        assert data[0].value == 'second'

    def test_sheet_not_found(self, tmp_path: Path, has_openpyxl: bool) -> None:
        """测试指定的工作表不存在"""
        if not has_openpyxl:
            pytest.skip("openpyxl not installed")

        from openpyxl import Workbook

        xlsx_file = tmp_path / "data.xlsx"
        wb = Workbook()
        wb.active.title = 'Sheet1'
        wb.save(str(xlsx_file))

        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class Data(Base):
            __tablename__ = 'data'
            id = Column(int, primary_key=True)

        with pytest.raises(ValueError) as exc_info:
            load_table(Data, str(xlsx_file), sheet_name='NonExistent')

        assert "not found" in str(exc_info.value)

    def test_excel_type_conversion(self, tmp_path: Path, has_openpyxl: bool) -> None:
        """测试 Excel 原生类型转换"""
        if not has_openpyxl:
            pytest.skip("openpyxl not installed")

        from openpyxl import Workbook
        from datetime import datetime

        xlsx_file = tmp_path / "data.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['id', 'score', 'created'])
        ws.append([1, 95.5, datetime(2024, 1, 15, 10, 30, 0)])
        wb.save(str(xlsx_file))

        db = Storage(in_memory=True)
        Base: Type[PureBaseModel] = declarative_base(db)

        class Data(Base):
            __tablename__ = 'data'
            id = Column(int, primary_key=True)
            score = Column(float)
            created = Column(datetime)

        data = load_table(Data, str(xlsx_file))

        assert len(data) == 1
        assert data[0].id == 1
        assert data[0].score == 95.5
        assert isinstance(data[0].created, datetime)
        assert data[0].created.year == 2024
